#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
模块名称：multi_thread_test
模块描述：提供基于多线程对doubao模型进行大规模数据测试的能力

作者：Albert Liu
创建日期：2024-07-11
版本：2.0

依赖：
    - openpyxl (安装方法: `pip3 install openpyxl`)
    - volcengine.maas.v2 (安装方法: `pip3 install volcengine`)
    - datasets (安装方法: `pip3 install datasets`)

用法示例：
    使用方式：
    1. 按测试数据集的数据结构，修改数据清洗方法
    - construct_test_data：读取test_data
        - construct_test_data_by_json: 读取json数据
        - construct_test_data_by_huggingface: 读取huggingface数据
    - construct_system_prompt：读取system_prompt文件，默认读取同目录下system_prompt.txt文件
    - construct_user_prompt：基于test_data中的数据结构，组装出user_prompt
    - result_process：对于doubao的response进行数据清洗
    - write_excel:将doubao返回内容写入文件，column 1默认写入当前数据在test_data中的index; column 2默认写入doubao返回的role为assistant的content

    2. 选择以下4种任务中，任选其一执行任务
    - 任务1，单线程：5个样本，并行跑，index含0不含5
    - 任务2，并行平均划分：200个样本到5个thread
    - 任务3，index划分并行：15个样本，分别划分到3个thread中，每个thread分别负责0-3、4-10、11-15
    - 任务4，串行点查：点查只查询index为40,55,100的三个值

    3. 结果输出
    结果都生成在SUB_PATH路径下，一般会生成3类文件
    - data{xx}-{xx}.xlsx: 多个线程运行后生成的临时文件
    - combined_output.xlsx 融合多个线程运行结果的最终输出数据
    - log.log: 业务运行日志
    - num.txt: 多次重时候，仍失败的任务index（test_data中的index）

    ☆注意：
    1. 建议只修改全局参数 及 需要修改的数据清洗方法，其他方法一般不需要修改
    2. 建议修改完各类construct_方法后，先用任务1跑一遍，无bug再起多线程并发跑数据
    3. 核心代码逻辑在basic_run, 其他如run_average_parallel何run_index_parallel都是对basic_run的多线程封装
"""
import functools
import json
import logging
import os
import threading
import time
import traceback
import warnings

import openpyxl
from datasets import load_dataset
from openpyxl.reader.excel import load_workbook
from openpyxl.workbook import Workbook
from volcengine.maas.v2 import MaasService


##################################### 完全不用看 ########################################

# 废弃方法装饰器
def deprecated(reason):
    """
    标注废弃方法的装饰器。
    :param reason: 废弃原因的描述。
    :return:
    """

    def decorator(func):
        @functools.wraps(func)
        def wrapper(*args, **kwargs):
            warnings.warn(
                f"Call to deprecated function {func.__name__} ({reason}).",
                category=DeprecationWarning,
                stacklevel=2
            )
            return func(*args, **kwargs)

        return wrapper

    return decorator


# 合并txt文件tool method
@deprecated("This function will be removed in future versions.")
def __merge_files(input_files, output_file):
    """
    将input_files中的多个文件，合成一个output_file文件
    :param input_files: 待合成文件的path_list
    :param output_file: 合成结果文件的path
    :return: None
    """
    with open(output_file, 'w', encoding='utf-8') as outfile:
        for file in input_files:
            with open(file, 'r', encoding='utf-8') as infile:
                content = infile.read()
                outfile.write(content)


# 合并文件业务逻辑
@deprecated("This function will be removed in future versions.")
def __combine_file(tag_list):
    """
    将多个文件合并，合并依据是tag_list。合并内容包括num.txt和combined_output.xlsx
    :param tag_list: index分割节点list，比如合并3个文件data0-5.xlsx，data6-7.xlsx，data8-10.xlsx，值为[[0, 6], [6, 8], [8, 10]]
    :return: None
    """
    # 文件目录
    output_num = os.path.join(SUB_PATH, "num.txt")
    output_excel = os.path.join(SUB_PATH, "combined_output.xlsx")

    # Excel
    # 创建一个新的工作簿
    combined_wb = Workbook()
    combined_ws = combined_wb.active
    # 依次读取剩余的Excel文件并追加其内容到新的工作簿
    for file in tag_list:
        try:
            wb = load_workbook(__tag_path("data", file[0], file[1], ".xlsx"))
            ws = wb.active
            for row in ws.iter_rows(values_only=True):
                combined_ws.append(row)
        except Exception as e:
            logger = logging.getLogger(f"__main__")
            logger.error("======== message ==========")
            logger.error(str(e) + "\n")
            logger.error(traceback.format_exc() + "\n")
            continue
    # 将合并后的数据保存到新的Excel文件
    combined_wb.save(output_excel)

    # num
    num_list = []
    for file in tag_list:
        # num.txt可能不存在
        num_path = __tag_path("num", file[0], file[1], ".txt")
        if os.path.exists(num_path):
            num_list.append(num_path)
    # 合并
    if len(num_list) != 0:
        __merge_files(num_list, output_num)


# 生成带tag的文件名
@deprecated("This function will be removed in future versions.")
def __tag_path(name, start, end, suffix):
    """
    为中间过程临时文件生成文件命名
    :param name: 真实文件名，如log
    :param start: 当前thread负责的start index
    :param end: 当前thread负责的end index
    :param suffix: 后缀，如.log
    :return:
    """
    return os.path.join(SUB_PATH, f"{name}{start}-{end - 1}{suffix}")


# 构造system_prompt和user_prompt构建request
def __construct_req(system_prompt, user_prompt):
    """
    基于system_prompt和user_prompt，以及PARAMETERS构建API接口要求的req数据结构, 一般不需要变更
    :param system_prompt:
    :param user_prompt:
    :return req: 组装后的request
    """
    req = {}
    req["parameters"] = PARAMETERS
    req["messages"] = [
        {
            "role": "system",
            "content": system_prompt
        }, {
            "role": "user",
            "content": user_prompt
        }
    ]
    return req


# 根据resp解析出返回content
def __construct_message(resp):
    """
    对doubao api接口返回的response结构进行解析，拿到所需内容
    :param resp: http返回结构
    :return message: 解析后的message结构如下
        "message": {
            "role": "xxx",
            "content": "xxx"
        }
    """
    message = None
    if resp is not None:
        message = resp['choices'][0]['message']
    return message


# doubao调用
def doubao(system_prompt, user_prompt):
    """
    向豆包发起请求，request的组装和response的后处理在此完成
    :param system_prompt:
    :param user_prompt:
    :return message:
    message数据结构，解析后的message结构如下
    "message": {
        "role": "xxx",
        "content": "xxx"
    }
    """
    # 如果调用的时候，遇到Read time out问题，可以尝试调大connection_timeout和socket_timeout
    maas = MaasService('maas-api.ml-platform-cn-beijing.volces.com', 'cn-beijing',
                       connection_timeout=600,
                       socket_timeout=600)
    maas.set_ak(VOLC_ACCESSKEY)
    maas.set_sk(VOLC_SECRETKEY)

    # 构建请求
    req = __construct_req(system_prompt, user_prompt)
    # chat接口调用
    resp = maas.chat(ENDPIOINT_ID, req)
    # resp结果处理
    content = __construct_message(resp)
    return content


##################################### 从这里开始看 ########################################


# test_data读取
def construct_test_data_by_json(file_path):
    """
    读取测试集，当测试集为json list文件时，可自定义

    :param file_path: json文件位置
    :return data: 返回一定是1个list数据结构，便于后续循环处理
    """
    # 打开 JSON 文件
    with open(file_path, "r", encoding='utf-8') as file:
        # 解析 JSON 文件内容
        data = json.load(file)
    return data


# test_data读取
def construct_test_data_by_huggingface(file_path):
    """
    读取测试集，当测试集为huggingface的dataset文件时，可自定义
    :param file_path: dataset文件位置
    :return dataset_list: 返回一定是1个list数据结构，便于后续循环处理
    """
    ## Load test data
    dataset = load_dataset('json', data_files=file_path)
    ## 提取测试集
    test_set = dataset['train']
    dataset_dict = test_set.to_dict()
    # dic to list
    dataset_list = [dict(zip(dataset_dict, t)) for t in zip(*dataset_dict.values())]
    return dataset_list


# 构建system_prompt
def construct_system_prompt(file_path):
    """
    构建system_prompt，默认读取当前目录下的system_prompt.txt文件

    :param file_path: system_prompt文件路径
    :return system_prompt:
    """
    try:
        with open(file_path, 'r', encoding='utf-8') as file:
            system_prompt = file.read()
        return system_prompt
    except FileNotFoundError:
        print("文件未找到")


# 构建user_prompt
def construct_user_prompt(item):
    """
    使用测试集中的元素，组装user_prompt，完全自定义user_prompt组装逻辑

    :param item: 对应test_data_list中某index的一个元素
    :return user_prompt:
    """
    problem = item["problem"]
    option_list = item["answer_option_list"]
    option_list = [str(arr) for arr in option_list]
    option_str = "\n".join(option_list)
    user_prompt = ("## 问题描述 ##" + "\n" +
                   problem + "\n" +
                   "## 问题选项 ##" + "\n" +
                   option_str)
    return user_prompt


# 创建excel
def construct_excel(excel_path):
    """
    构建存放测试结果的excel，指定目录和拿到句柄，实际数据处理在write_excel方法中

    :param excel_path: 结果excel文件目录
    :return excel_info: 当前excel的写入句柄等信息
    :return start_row: 当前excel待写入行数
    """
    sheet_name = "data"
    # 表格处理
    if os.path.exists(excel_path):
        # 数据追加
        workbook = openpyxl.load_workbook(excel_path)
        worksheet = workbook[sheet_name]
        worksheet._current_row = worksheet.max_row
    else:
        # 新创sheet及file
        workbook = Workbook()
        worksheet = workbook.create_sheet(sheet_name, 0)
    max_row = worksheet.max_row

    start_row = max_row + 1
    excel_info = {
        "excel_path": excel_path,
        "worksheet": worksheet,
        "workbook": workbook,
    }
    return excel_info, start_row


# doubao返回结果后处理
def result_process(index, system_prompt, user_prompt, response_data, raw_data):
    """
    基于doubao返回的message进行后处理，逻辑完全自定义，处理完成后的数据为1个待写入excel的list

    :param index: 当前result在test_data list中的index
    :param system_prompt:
    :param user_prompt:
    :param response_data: doubao返回的response数据
    :param raw_data: test_data 中某原始结果数据
    :return result: 1个待写入excel的list
    """

    # 结果处理，清除转义字符，json解析返回结果
    resp_text = response_data["content"]
    resp_text = resp_text.replace("\\", "")
    resp_json = json.loads(resp_text)

    # 结果组装
    result = [index, system_prompt, user_prompt,
              resp_json["answer"], resp_json["analysis"],
              raw_data["answer_value"], raw_data["answer_analysis"][0]]
    return result


# 写入excel
def write_excel(excel_dic, result_list, row_num):
    """
    将result_list结果数据写入到excel的第row_num行中

    :param excel_dic: 待写入的excel文件信息
    :param result_list: 结果list
    :param row_num: 待写入excel的行号
    :return None:
    """
    # doubao返回结果处理
    for index, item in enumerate(result_list):
        excel_dic["worksheet"].cell(row=row_num, column=index + 1, value=item)
    # 保存结果
    excel_dic["workbook"].save(excel_dic["excel_path"])


##################################### 并发调用逻辑 ########################################


# 核心调用逻辑
def basic_run(start, end, index_list, thread_num, excel_dic, row_num):
    """
    单线程计算逻辑，基础运算逻辑如下
    1. 基于test_data_list循环遍历，请求doubao给出请求结果
    2. 添加多次重试机制，当某个test_data测试报错，等待WAIT_TIME秒后，重试ATTEMPT_TIME次
    3. 3个重要文件记录数据
    - log.log文件记录日志
    - data.xlsx文件保存结果数据
    - num文件记录失败test_data的index（num文件仅当有ATTEMPT_TIME次尝试仍失败后才出现）

    :param start: 循环开始index
    :param end: 循环结束index
    :param index_list: 待测试的index编号。当index_list不为None时，start和end失效，只遍历index_list中的数据
    :param thread_num: 当前thread的编号
    :param excel_dic: excel数据信息
    :param row_num: 初始写入行号
    :return None:
    """
    logger = logging.getLogger(f"__thread {thread_num}__")

    # 数据读取
    system_prompt = construct_system_prompt(SYSTEM_PROMPT_PATH)
    test_data_list = construct_test_data_by_huggingface(TEST_DATA_PATH)

    # 遍历test_data
    for index, item in enumerate(test_data_list):
        # 循环控制-点查/顺序查
        if index_list is not None:
            # 根据index_list点查
            if index not in index_list:
                continue
        else:
            # 根据start和end顺序查
            if index >= end:
                break
            elif index < start:
                continue

        for j in range(ATTEMPT_TIME):
            response_dic = None
            try:
                # 组装user_prompt
                user_prompt = construct_user_prompt(item)
                # 请求doubao
                message = doubao(system_prompt, user_prompt)
                # 结果后处理
                result_list = result_process(index, system_prompt, user_prompt, message, item)
                # 结果写入excel
                with EXCEL_LOCK:
                    write_excel(excel_dic, result_list, row_num)
                # 换行
                row_num += 1
                # 打印日志
                logger.info(f"the {index} problem solved success！")
                break
            except Exception as e:
                with LOG_LOCK:
                    # 输出错误提示
                    logger.error(f"the {index} problem {j + 1} attempt failed！")
                    if j >= ATTEMPT_TIME - 1:
                        # 记录错误日志
                        logger.error(f"======message======: \n")
                        logger.error(str(e) + "\n")
                        logger.error(traceback.format_exc() + "\n")
                        if response_dic is not None:
                            logger.error(f"======content======: \n")
                            logger.error(response_dic["message"])
                # 记录错误序号
                if j >= ATTEMPT_TIME - 1:
                    with NUM_LOCK:
                        with open(os.path.join(SUB_PATH, "num.txt"), "a") as num_file:
                            num_file.write(f"{index}, ")
                time.sleep(WAIT_TIME)


# 数据集均分，并发
def run_average_parallel(total, thread_total, excel_dic, start_row):
    """
    当测试机可以基于多线程进行均分计算，可使用此方法。此方法基于basic_run方法构建

    :param total: 一共多少个测试用例
    :param thread_total: 要用几个线程跑
    :param excel_dic: excel数据信息
    :param start_row: 初始写入行号
    :return None:
    """
    logger = logging.getLogger(f"__main__")
    logger.info("开始计算")

    start = 0
    step = total // thread_total

    # thread pool
    threads = []
    for thread_num in range(thread_total):
        index_min = thread_num * step + start
        index_max = (thread_num + 1) * step + start

        start_row_index = start_row + (step * thread_num)
        # 并行
        t = threading.Thread(target=basic_run,
                             args=(index_min, index_max, None, thread_num, excel_dic, start_row_index))
        threads.append(t)
        t.start()
    for t in threads:
        t.join()
    logger = logging.getLogger(f"__main__")
    logger.info("计算完成")


# 根据index划分数据集，并发
def run_index_parallel(index_list, excel_dic, start_row):
    """
    点查，对index_list中的测试用例进行顺序计算，一般用于多线程计算后，对部分失败用例进行一次计算

    :param index_list: 待计算index列表
    :param excel_dic: excel数据信息
    :param start_row: 初始写入行号
    :return: None
    """
    logger = logging.getLogger(f"__main__")
    logger.info("开始计算")

    # thread pool
    threads = []
    for thread_num in range(len(index_list) - 1):
        index_min = index_list[thread_num]
        index_max = index_list[thread_num + 1]

        # 并行
        t = threading.Thread(target=basic_run,
                             args=(index_min, index_max, None, thread_num, excel_dic, start_row))
        threads.append(t)
        t.start()

        step = index_max - index_min
        start_row += step

    for t in threads:
        t.join()
    logger = logging.getLogger(f"__main__")
    logger.info("计算完成")


if __name__ == '__main__':

    ##################################### 全局参数 ########################################
    # doubao参数设置
    # 凭证信息
    VOLC_ACCESSKEY = ""
    VOLC_SECRETKEY = ""
    ENDPIOINT_ID = ""
    # 模型参数设定
    PARAMETERS = {
        "max_new_tokens": 1000,  # 输出文本的最大tokens限制
        "min_new_tokens": 1,  # 输出文本的最小tokens限制
        "temperature": 1,  # 用于控制生成文本的随机性和创造性，Temperature值越大随机性越大，取值范围0~1
        "top_p": 0.7,  # 用于控制输出tokens的多样性，TopP值越大输出的tokens类型越丰富，取值范围0~1
        "top_k": 0,  # 选择预测值最大的k个token进行采样，取值范围0-1000，0表示不生效
        "max_prompt_tokens": 3000,  # 最大输入 token 数，如果给出的 prompt_gen 的 token 长度超过此限制，取最后 max_prompt_tokens 个 token 输入模型。
        "repetition_penalty": 1.1  # 重复token输出的惩罚项
    }

    # 声明各文件位置
    # 临时文件存储目录
    SUB_PATH = "./tmp"
    if not os.path.exists(SUB_PATH):
        os.makedirs(SUB_PATH)
    # system prompt
    SYSTEM_PROMPT_PATH = 'system_prompt.txt'
    # 测试数据
    TEST_DATA_PATH = './data/TAL-SCQ5K/ch_single_choice_constructed_5K/ch_single_choice_train_3K.jsonl'

    # thread基础设置
    # 线程失败重试次数
    ATTEMPT_TIME = 3
    # 线程失败重试间隔
    WAIT_TIME = 5

    # 文件锁
    EXCEL_LOCK = threading.Lock()
    LOG_LOCK = threading.Lock()
    NUM_LOCK = threading.Lock()

    # 配置日志记录器
    logging.basicConfig(
        level=logging.INFO,  # 设置日志级别
        format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',  # 设置日志格式
        handlers=[
            logging.FileHandler(os.path.join(SUB_PATH, "log.log")),  # 将日志写入文件
            logging.StreamHandler()  # 将日志输出到控制台
        ]
    )

    # excel文件存放位置
    excel_path = os.path.join(SUB_PATH, "data.xlsx")
    excel_dic, start_row = construct_excel(excel_path)

    ################################# 以下任务任选其一 ######################################

    # # 任务1，单线程：5个样本，并行跑，index含0不含5
    basic_run(0, 5, None, 0, excel_dic, start_row)

    # 任务2，并行平均划分：200个样本到5个thread
    run_average_parallel(20, 10, excel_dic, start_row)

    # 任务3，index划分并行：15个样本，分别划分到3个thread中，每个thread分别负责0-3、4-10、11-15
    # sepreate_index_list = [0, 4, 11, 15]
    sepreate_index_list = [0, 6, 8, 10, 13, 20, 30]
    run_index_parallel(sepreate_index_list, excel_dic, start_row)

    # 任务4，串行点查：点查只查询index为40,55,100的三个值
    point_index_list = [0, 40, 55, 110]
    basic_run(0, 400, point_index_list, 0, excel_dic, start_row)
